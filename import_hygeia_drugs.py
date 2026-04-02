"""
Import Hygeia DrugGenericName Excel data into MySQL table drug_generic_names.

Requirements:
- openpyxl
- mysql-connector-python
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Sequence, Set, Tuple

import mysql.connector
from openpyxl import load_workbook


DB_CONFIG = {
    "host": "100.94.208.11",
    "user": "admin",
    "password": "123456",
    "database": "syntropic_rx",
}

SHEET_NAME = "DrugGenericName"
BATCH_SIZE = 500

SOURCE_PATH_CANDIDATES = (
    Path("/Users/CYUT/Documents/GitHub/syntropic-rx/DrugGenericName.xlsx"),
    Path(__file__).resolve().parent / "DrugGenericName.xlsx",
)

SOURCE_TO_DB = {
    "Code": "code",
    "Name": "name",
    "IndicationNote": "indication_note",
    "IsAntibiotic": "is_antibiotic",
    "IsPregnancyLactation": "is_pregnancy_lactation",
    "IsPregnancyCategory": "is_pregnancy_category",
    "PregnancyCategoryTypeKey": "pregnancy_category_type_key",
    "DrugGroupKey": "drug_group_id",
    "IsDisabled": "is_disabled",
}

BOOLEAN_COLUMNS = {
    "IsAntibiotic",
    "IsPregnancyLactation",
    "IsPregnancyCategory",
    "IsDisabled",
}

INT_COLUMNS = {
    "PregnancyCategoryTypeKey",
    "DrugGroupKey",
}


def resolve_source_path() -> Path:
    for path in SOURCE_PATH_CANDIDATES:
        if path.exists() and path.is_file():
            return path
    raise FileNotFoundError(
        "Could not find DrugGenericName.xlsx. Checked: "
        + ", ".join(str(p) for p in SOURCE_PATH_CANDIDATES)
    )


def normalize_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 1 if int(value) != 0 else 0
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true", "t", "yes", "y"}:
            return 1
        if normalized in {"0", "false", "f", "no", "n", ""}:
            return 0
    raise ValueError(f"Unsupported boolean value: {value!r}")


def normalize_int(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        raw = value.strip()
        if raw == "":
            return None
        return int(float(raw))
    raise ValueError(f"Unsupported integer value: {value!r}")


def normalize_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def read_rows_from_excel(source_path: Path) -> List[Tuple]:
    wb = load_workbook(filename=source_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {source_path}")

    ws = wb[SHEET_NAME]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return []

    header_map: Dict[str, int] = {}
    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        header_map[str(cell_value).strip()] = idx

    missing = [col for col in SOURCE_TO_DB if col not in header_map]
    if missing:
        wb.close()
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows: List[Tuple] = []
    source_columns: Sequence[str] = list(SOURCE_TO_DB.keys())

    for row_values in ws.iter_rows(min_row=2, values_only=True):
        values_by_col = {}
        is_empty_row = True

        for source_col in source_columns:
            value = row_values[header_map[source_col]]

            if source_col in BOOLEAN_COLUMNS:
                value = normalize_bool(value)
            elif source_col in INT_COLUMNS:
                value = normalize_int(value)
            else:
                value = normalize_text(value)

            if value is not None:
                is_empty_row = False
            values_by_col[source_col] = value

        if is_empty_row:
            continue

        rows.append(tuple(values_by_col[col] for col in source_columns))

    wb.close()
    return rows


def chunked(items: List[Tuple], size: int) -> Sequence[List[Tuple]]:
    for i in range(0, len(items), size):
        yield items[i : i + size]


def load_valid_drug_group_ids(cursor) -> Set[int]:
    cursor.execute("SELECT id FROM drug_groups")
    return {int(row[0]) for row in cursor.fetchall()}


def sanitize_drug_group_ids(rows: List[Tuple], valid_group_ids: Set[int]) -> Tuple[List[Tuple], int]:
    if not rows:
        return rows, 0

    source_columns = list(SOURCE_TO_DB.keys())
    drug_group_idx = source_columns.index("DrugGroupKey")
    fixed_rows: List[Tuple] = []
    nullified_count = 0

    for row in rows:
        values = list(row)
        group_id = values[drug_group_idx]

        if group_id is not None and group_id not in valid_group_ids:
            values[drug_group_idx] = None
            nullified_count += 1

        fixed_rows.append(tuple(values))

    return fixed_rows, nullified_count


def main() -> None:
    source_path = resolve_source_path()
    rows = read_rows_from_excel(source_path)

    source_columns = list(SOURCE_TO_DB.keys())
    db_columns = [SOURCE_TO_DB[col] for col in source_columns]
    placeholders = ", ".join(["%s"] * len(db_columns))
    sql = (
        f"INSERT IGNORE INTO drug_generic_names ({', '.join(db_columns)}) "
        f"VALUES ({placeholders})"
    )

    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    valid_group_ids = load_valid_drug_group_ids(cursor)
    rows, nullified_group_count = sanitize_drug_group_ids(rows, valid_group_ids)

    inserted_count = 0
    try:
        for batch in chunked(rows, BATCH_SIZE):
            cursor.executemany(sql, batch)
            conn.commit()
            inserted_count += max(cursor.rowcount, 0)
    finally:
        cursor.close()
        conn.close()

    print("=" * 60)
    print("Hygeia Drug Import Summary")
    print("=" * 60)
    print(f"Source file         : {source_path}")
    print(f"Rows read from Excel: {len(rows):,}")
    print(f"DrugGroupKey -> NULL: {nullified_group_count:,}")
    print(f"Records imported    : {inserted_count:,}")


if __name__ == "__main__":
    main()
